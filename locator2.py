from PIL import ImageGrab
import cv2
from lib import CTDT
import time
import os
import numpy as np
from os.path import isfile, join
from os import listdir
import openpyxl
from openpyxl import Workbook, load_workbook, worksheet
import ctypes

user32 = ctypes.windll.user32
user32.SetProcessDPIAware()

wb: Workbook = load_workbook("data.xlsx")
ws: worksheet = wb["Templates"]

threshold = 0.9

start_index = 1
end_index = 38
index = start_index

while index <= end_index:
    template_number = "{0}".format(str(index).zfill(3))

    image_screen = cv2.imread(os.path.join("templates_original", template_number + "f.png"))
    image_template = cv2.imread(os.path.join("templates_original", template_number + ".png"), 0)
    width, height = image_template.shape[::-1]

    image_rgb = np.array(image_screen)
    image_gray = cv2.cvtColor(image_rgb, cv2.COLOR_BGR2GRAY)
    res = cv2.matchTemplate(image_gray, image_template, cv2.TM_CCOEFF_NORMED)

    loc = np.where(res >= threshold)

    if len(loc[0]) == 0 & len(loc[1]) == 0:
        print("Not found")
    else:

        print(
            "########################################### {0} ###################################################".format(
                template_number))

        start_x = loc[1][0]
        ws.cell(index + 1, 2, start_x - 5)

        start_y = loc[0][0]
        ws.cell(index + 1, 3, start_y - 5)

        end_x = start_x + width
        ws.cell(index + 1, 4, end_x + 5)

        end_y = start_y + height
        ws.cell(index + 1, 5, end_y + 5)

        print("Start X : {0}".format(start_x - 5))
        print("Start Y : {0}".format(start_y - 5))

        print("End X : {0}".format(end_x + 5))
        print("End Y : {0}".format(end_y + 5))

    index += 1

wb.save("data.xlsx")
