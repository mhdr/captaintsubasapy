from PIL import ImageGrab
import cv2
from openpyxl.worksheet.worksheet import Worksheet

from lib import CTDT
import time
import os
import numpy as np
import ctypes
from openpyxl import Workbook, load_workbook

user32 = ctypes.windll.user32
user32.SetProcessDPIAware()

threshold = 0.97

# file name like : 031f.jpg
# Column A in excel
index_fimage = 65

# template number like : 007.jpg
# Column F in excel
index_template = 65

# first match = 0
match = 0

# Window name in which image is displayed
window_name = 'Image'

template_number = "{0}".format(str(index_template).zfill(3))
f_image_number = "{0}".format(str(index_fimage).zfill(3))

image_screen = cv2.imread(os.path.join("templates_original", f_image_number + "f.png"))
image_template = cv2.imread(os.path.join("templates_original", template_number + ".png"), 0)
width, height = image_template.shape[::-1]

image_rgb = np.array(image_screen)
image_gray = cv2.cvtColor(image_rgb, cv2.COLOR_BGR2GRAY)
res = cv2.matchTemplate(image_gray, image_template, cv2.TM_CCOEFF_NORMED)

loc = np.where(res >= threshold)

if len(loc[0]) == 0 & len(loc[1]) == 0:
    print("Not found")
else:

    start_x = loc[1][match]
    start_y = loc[0][match]

    end_x = start_x + width
    end_y = start_y + height

    print("Start X : {0}".format(start_x - 5))
    print("Start Y : {0}".format(start_y - 5))

    print("End X : {0}".format(end_x + 5))
    print("End Y : {0}".format(end_y + 5))

    print("Width : {0}".format(end_x - start_x))
    print("Height : {0}".format(end_y - start_y))

    start_point = (start_x, start_y)
    end_point = (end_x, end_y)

    # Red color in BGR
    color = (0, 0, 255)

    # Line thickness of 2 px
    thickness = 5

    # Using cv2.rectangle() method
    # Draw a rectangle with blue line borders of thickness of 2 px
    image = cv2.rectangle(image_screen, start_point, end_point, color, thickness)

    cv2.namedWindow(window_name, cv2.WND_PROP_FULLSCREEN)
    cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
    # Displaying the image
    cv2.imshow(window_name, image)

    # save to data file
    wb: Workbook = load_workbook(filename="data.xlsx")
    ws: Worksheet = wb["Templates"]
    end_row = ws.max_row
    # start after header
    start_row = 2
    row_index = start_row
    while row_index <= end_row:

        if ws["A" + str(row_index)].value == template_number:
            ws.cell(row_index, 2, start_x - 5)
            ws.cell(row_index, 3, start_y - 5)
            ws.cell(row_index, 4, end_x + 5)
            ws.cell(row_index, 5, end_y + 5)
            break
        row_index += 1

    wb.save("data.xlsx")
    CTDT.convert_templates_to_jpeg()
    cv2.waitKey(0)
