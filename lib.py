import collections
import time
from datetime import datetime
import cv2
import pyautogui
from os import listdir
from os.path import isfile, join
from PIL import ImageGrab
from openpyxl import Workbook, worksheet, load_workbook
import os
from PIL import Image
import numpy as np
from dataclasses import dataclass
from typing import Dict
import datetime
import configparser
import ctypes
import shutil
from openpyxl.worksheet.worksheet import Worksheet
import pytesseract


#######################################################################################################################

@dataclass()
class Box:
    left: int
    top: int
    width: int
    height: int


#######################################################################################################################

@dataclass()
class TemplateProperties:
    template_number: str
    image: any
    image_width: int
    image_height: int
    region_start_x: int
    region_start_y: int
    region_end_x: int
    region_end_y: int
    date_seen: datetime.datetime


#######################################################################################################################

@dataclass()
class LocationProperties:
    location_number: str
    x: int
    y: int


#######################################################################################################################

class Cache:
    __instance = None
    templates: Dict[str, TemplateProperties] = {}
    locations: Dict[str, LocationProperties] = {}

    @staticmethod
    def get_instance():
        """
        static access method
        :return: Template
        """
        if Cache.__instance is None:
            Cache()

        return Cache.__instance

    def __init__(self):
        """
        Virtually private constructor
        """
        Cache.__instance = self


#######################################################################################################################

class Config:
    __instance = None
    mode: int
    difficulty: int
    sleep: float
    prevent_screen_off: int
    energy_recovery: int
    wait_energy_recovery: int
    wait_telegram_msg_energy_recovery: int
    play_match_with_skip_ticket_button: int
    global_shared_play_enabled: int
    wait_finish_ad: int
    wait_ad_view_interrupted: int
    max_count_now_loading: int
    max_count_preparing: int
    max_count_sharing: int
    min_recovery_ball: int

    telegram_token: str
    telegram_chatid: int
    telegram_disabled: int
    telegram_token2: str
    telegram_chatid2: int

    wait_after_member1_join: int
    wait_after_member2_join: int
    wait_after_member3_join: int

    # number of energy left for energy recovery
    energy_ad_left: int

    wait_before_go_home: int

    @staticmethod
    def get_instance():
        """
        static access method
        :return: Template
        """
        if Config.__instance is None:
            config = configparser.ConfigParser()
            config.read('config.ini')

            Config.sleep = float(config["General"]["Sleep"])
            Config.prevent_screen_off = int(config["General"]["PreventScreenOff"])

            Config.mode = int(config["Game"]["Mode"])
            Config.difficulty = int(config["Game"]["Difficulty"])
            Config.energy_recovery = int(config["Game"]["EnergyRecovery"])
            Config.wait_energy_recovery = int(config["General"]["WaitForEnergyRecovery"])
            Config.wait_telegram_msg_energy_recovery = int(config["General"]["WaitTelegramMsgEnergyRecovery"])
            Config.play_match_with_skip_ticket_button = int(config["Game"]["PlayMatchWithSkipTicketButton"])
            Config.global_shared_play_enabled = int(config["Game"]["GlobalSharedPlayEnabled"])
            Config.wait_finish_ad = int(config["General"]["WaitFinishAd"])
            Config.wait_ad_view_interrupted = int(config["General"]["WaitAfterAdViewInterrupted"])
            Config.max_count_now_loading = int(config["General"]["MaxCountNowLoading"])
            Config.max_count_preparing = int(config["General"]["MaxCountPreparing"])
            Config.max_count_sharing = int(config["General"]["MaxCountSearching"])
            Config.min_recovery_ball = int(config["General"]["MinRecoveryBall"])
            Config.wait_after_member1_join = int(config["General"]["WaitAfterMember1Joined"])
            Config.wait_after_member2_join = int(config["General"]["WaitAfterMember2Joined"])
            Config.wait_after_member3_join = int(config["General"]["WaitAfterMember3Joined"])
            Config.energy_ad_left = int(config["Game"]["EnergyAdLeft"])
            Config.wait_before_go_home = int(config["General"]["WaitBeforeGoHome"])

            Config.telegram_token = str(config["Telegram"]["Token"])
            Config.telegram_chatid = int(config["Telegram"]["ChatId"])
            Config.telegram_disabled = int(config["Telegram"]["Disabled"])
            Config.telegram_token2 = str(config["Telegram"]["Token2"])
            Config.telegram_chatid2 = int(config["Telegram"]["ChatId2"])

            Config()

        return Config.__instance

    def get_text_mode(self, mode: int = 0):
        # Story Solo = 1
        # Event Solo = 2
        # Solo = 3 ( general use )
        # Club Shared = 4
        # Club Join = 5
        # Global Shared = 6
        # Global Recruit = 7
        # Evolve Player = 8
        # Farm Story Mode = 9

        #  if 0 load from instance
        if mode == 0:
            mode = self.mode

        msg = ""

        if mode == 1:
            msg = "Mode 1 : Story Solo"
        elif mode == 2:
            msg = "Mode 2 : Event Solo"
        elif mode == 3:
            msg = "Mode 3 : Solo"
        elif mode == 4:
            msg = "Mode 4 : Club Shared"
        elif mode == 5:
            msg = "Mode 5 : Club Join"
        elif mode == 6:
            msg = "Mode 6 : Global Shared"
        elif mode == 7:
            msg = "Mode 7 : Global Recruit"
        elif mode == 8:
            msg = "Mode 8 : Evolve Player"
        elif mode == 9:
            msg = "Mode 9 : Farm Story Mode"

        return msg

    def get_text_difficulty(self, difficulty: int = 0):

        # Normal = 1
        # Hard = 2
        # Very Hard = 3
        # Extreme = 4

        #  if 0 load from instance
        if difficulty == 0:
            difficulty = self.difficulty

        msg = ""

        if difficulty == 1:
            msg = "Difficulty 1 : Normal"
        elif difficulty == 2:
            msg = "Difficulty 2 : Hard"
        elif difficulty == 3:
            msg = "Difficulty 3 : Very Hard"
        elif difficulty == 4:
            msg = "Difficulty 4 : Extreme"

        return msg

    def __init__(self):
        """
        Virtually private constructor
        """
        Config.__instance = self


#######################################################################################################################

class LocateResult:
    template: TemplateProperties = None
    position: Box = None

    def __init__(self, temp: TemplateProperties = None, pos: Box = None):
        self.template = temp
        self.position = pos

    def click(self, wait: float = 2, delay=0, full_screen=False) -> bool:
        if self.position is not None:

            if full_screen is False:
                center_x = self.template.region_start_x + self.position.left + self.template.image_width / 2
                center_y = self.template.region_start_y + self.position.top + self.template.image_height / 2
            else:
                center_x = self.position.left + self.template.image_width / 2
                center_y = self.position.top + self.template.image_height / 2

            # pyautogui.moveTo(center_x, center_y)
            time.sleep(delay)
            pyautogui.click(center_x, center_y)
            pyautogui.FAILSAFE = False
            pyautogui.moveTo(0, 0)
            print("Click Template => {0} : {1}".format(self.template.template_number, datetime.datetime.now()))
            time.sleep(wait)
            return True

        return False

    def move_mouse(self):
        if self.position is not None:
            center_x = self.template.region_start_x + self.position.left + self.template.image_width / 2
            center_y = self.template.region_start_y + self.position.top + self.template.image_height / 2
            pyautogui.moveTo(center_x, center_y)

    def available(self):
        if self.position is not None:
            return True

        return False


#######################################################################################################################

class PointResult:
    location: LocationProperties = None

    def __init__(self, location=None):
        self.location = location

    def click(self, clicks: int = 1, interval: float = 0, wait: float = 2,
              delay: float = 0, duration=0.1) -> bool:
        if self.location is not None:
            time.sleep(delay)
            pyautogui.moveTo(self.location.x, self.location.y, duration)
            pyautogui.click(self.location.x, self.location.y, clicks=clicks, interval=interval)
            pyautogui.FAILSAFE = False
            pyautogui.moveTo(0, 0)
            time.sleep(wait)
            return True

        return False


#######################################################################################################################

class CTDT:

    @staticmethod
    def initialize():

        # make current process high priority
        # pid = win32api.GetCurrentProcessId()
        # handle = win32api.OpenProcess(win32con.PROCESS_ALL_ACCESS, True, pid)
        # win32process.SetPriorityClass(handle, win32process.HIGH_PRIORITY_CLASS)

        # make current process dpi aware
        user32 = ctypes.windll.user32
        user32.SetProcessDPIAware()
        pyautogui.FAILSAFE = False
        pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files (x86)\\Tesseract-OCR\\tesseract.exe'

    @staticmethod
    def initialize_cache():
        caches: Cache = Cache.get_instance()
        dir = "templates"

        wb: Workbook = load_workbook(filename="data.xlsx")
        ws: worksheet = wb["Templates"]
        ws2: worksheet = wb["Points"]

        end_row = ws.max_row
        # start after header
        start_row = 2
        row_index = start_row

        while row_index <= end_row:
            try:

                template_number: str = str(ws["A" + str(row_index)].value)
                start_x: int = int(ws["B" + str(row_index)].value)
                start_y: int = int(ws["C" + str(row_index)].value)
                end_x: int = int(ws["D" + str(row_index)].value)
                end_y: int = int(ws["E" + str(row_index)].value)
                # using separate file name help to use one image with multiple location
                # for most of the templates file name is the same as template number
                # but for some of them it is not
                template_file_name: str = str(ws["F" + str(row_index)].value)
                filename = join(dir, template_file_name + ".jpg")

                # Using 0 to read image in grayscale mode
                image = cv2.imread(filename, 0)
                width, height = image.shape[::-1]

                template_properties: TemplateProperties = TemplateProperties(template_number, image, width, height,
                                                                             start_x,
                                                                             start_y,
                                                                             end_x, end_y, None)

                caches.templates[template_number] = template_properties
                row_index += 1

            except Exception as ex:
                raise ex
                # breakpoint()

        end_row2 = ws2.max_row
        start_row2 = 2
        row_index2 = start_row2

        while row_index2 <= end_row2:
            location_number: str = str(ws2["A" + str(row_index2)].value)
            x: int = int(ws2["B" + str(row_index2)].value)
            y: int = int(ws2["C" + str(row_index2)].value)

            location_properties = LocationProperties(location_number, x, y)

            caches.locations[location_number] = location_properties
            row_index2 += 1

    @staticmethod
    def convert_templates_to_jpeg():
        dest_dir = "templates"
        src_dir = "templates_original"

        files_to_remove = [f for f in listdir(dest_dir) if isfile(join(dest_dir, f))]
        for fr in files_to_remove:
            os.remove(join(dest_dir, fr))

        files = [f for f in listdir(src_dir) if isfile(join(src_dir, f))]

        for file in files:
            filename = os.path.splitext(os.path.basename(file))[0]

            if filename.endswith("f"):
                continue

            image: Image.Image = Image.open(join(src_dir, file))
            image_rgb = image.convert("RGB")
            # image_rgb = image.convert("L")
            image_rgb.save(join(dest_dir, filename + ".jpg"), format='JPEG', quality=90)

    @staticmethod
    def point(location_number: str) -> PointResult:
        caches: Cache = Cache.get_instance()
        x = caches.locations[location_number].x
        y = caches.locations[location_number].y

        location: LocationProperties = LocationProperties(location_number, x, y)
        result: PointResult = PointResult(location)

        return result

    @staticmethod
    def template(template_number: str, threshold=0.9, full_screen=False) -> LocateResult:
        caches: Cache = Cache.get_instance()
        region_start_x = caches.templates[template_number].region_start_x
        region_start_y = caches.templates[template_number].region_start_y
        region_end_x = caches.templates[template_number].region_end_x
        region_end_y = caches.templates[template_number].region_end_y

        if full_screen is False:
            image_region = ImageGrab.grab(bbox=(region_start_x, region_start_y, region_end_x, region_end_y))
        else:
            image_region = ImageGrab.grab()

        image_template = caches.templates[template_number].image

        image_rgb = np.array(image_region)
        image_gray = cv2.cvtColor(image_rgb, cv2.COLOR_BGR2GRAY)
        res = cv2.matchTemplate(image_gray, image_template, cv2.TM_CCOEFF_NORMED)

        loc = np.where(res >= threshold)

        result: LocateResult = LocateResult()

        if len(loc[0]) == 0 & len(loc[1]) == 0:
            return result
        else:
            position = Box(loc[0][0], loc[1][0], caches.templates[template_number].image_width,
                           caches.templates[template_number].image_height)
            result = LocateResult(caches.templates[template_number], position)
            caches.templates[template_number].date_seen = datetime.datetime.now()

        return result

    @staticmethod
    def save_screenshot():
        try:
            user32 = ctypes.windll.user32
            user32.SetProcessDPIAware()
            image = ImageGrab.grab()
            image.save("screenshot.jpg")
        except Exception as ex:
            print(str(ex))

    @staticmethod
    def ocr_text(template_number: str, invert_colors=True) -> str:

        try:
            caches: Cache = Cache.get_instance()
            region_start_x = caches.templates[template_number].region_start_x
            region_start_y = caches.templates[template_number].region_start_y
            region_end_x = caches.templates[template_number].region_end_x
            region_end_y = caches.templates[template_number].region_end_y

            image_region = ImageGrab.grab(bbox=(region_start_x, region_start_y, region_end_x, region_end_y))
            image_rgb = np.array(image_region)

            image_gray = cv2.cvtColor(image_rgb, cv2.COLOR_BGR2GRAY)
            final_image = None

            if invert_colors == True:
                final_image = cv2.bitwise_not(image_gray)  # color of text is white so we should invert colors
            else:
                final_image = image_gray

            data = pytesseract.image_to_string(final_image, config='--psm 6')

            return data
        except:
            return ""

    @staticmethod
    def ocr_number(template_number: str, invert_colors=True) -> int:

        try:
            caches: Cache = Cache.get_instance()
            region_start_x = caches.templates[template_number].region_start_x
            region_start_y = caches.templates[template_number].region_start_y
            region_end_x = caches.templates[template_number].region_end_x
            region_end_y = caches.templates[template_number].region_end_y

            image_region = ImageGrab.grab(bbox=(region_start_x, region_start_y, region_end_x, region_end_y))
            image_rgb = np.array(image_region)

            image_gray = cv2.cvtColor(image_rgb, cv2.COLOR_BGR2GRAY)
            final_image = None

            if invert_colors == True:
                final_image = cv2.bitwise_not(image_gray)  # color of text is white so we should invert colors
            else:
                final_image = image_gray

            data = pytesseract.image_to_string(final_image, config='--psm 6')

            new_data = ""

            for d in data:
                if d.isnumeric():
                    new_data += d

            if new_data == "":
                new_data = "0"

            return int(new_data)
        except:
            return 0
