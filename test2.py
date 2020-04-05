from openpyxl import Workbook, worksheet, load_workbook

wb: Workbook = load_workbook(filename="data.xlsx")
ws: worksheet = wb.active

end_row = ws.max_row
# start after header
start_row = 2
row_index = start_row

while row_index <= end_row:
    print(ws["A" + str(row_index)].value)
    print(ws["B" + str(row_index)].value)
    print(ws["C" + str(row_index)].value)
    print(ws["D" + str(row_index)].value)
    print(ws["E" + str(row_index)].value)
    row_index += 1
